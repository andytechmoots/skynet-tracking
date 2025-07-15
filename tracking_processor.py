import pandas as pd
from datetime import timedelta
import os
import glob
from utils import unzip_file
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Column mapping
COLS = {
    'connote': 'Connote #',
    'manifest': 'Manifest Date',
    'AA': 'Date Released From Customs',
    'AB': 'Date Received from Customs Agent',
    'AC': 'Date Collected by Courier Provider',
    'AD': 'Arrived Hub Date',
    'AE': 'First OFD Date',
    'AF': 'POD Date',
    'driver': 'OFD Driver Name'
}

highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

def fill_between(df, col_prev, col_target, col_next, tracking_log, sheet_name, file_path):
    count = 0
    for i, row in df.iterrows():
        if pd.isna(row[col_target]):
            time_before = pd.to_datetime(row[col_prev], errors='coerce')
            time_after = pd.to_datetime(row[col_next], errors='coerce')
            if pd.notna(time_before) and pd.notna(time_after) and time_before < time_after:
                estimate = time_before + (time_after - time_before) / 2
                df.at[i, col_target] = estimate
                count += 1
                tracking_log.append({
                    'Connote #': row[COLS['connote']],
                    'Updated Column': col_target,
                    'Value': estimate,
                    'Sheet': sheet_name,
                    'Source File': os.path.basename(file_path),
                    'Row Index': i
                })
    return count

def process_skynet_reports(input_folder, output_folder):
    log = []
    tracking_log = []

    for file_path in glob.glob(f"{input_folder}/*.xls"):
        print(f"🔍 Reading {file_path}")
        try:
            # Try loading as true Excel
            excel = pd.ExcelFile(file_path, engine='xlrd')
            sheet_names = excel.sheet_names
            read_sheet = lambda sheet: pd.read_excel(file_path, sheet_name=sheet, engine='xlrd')
        except Exception as e:
            print(f"⚠️ Fallback to HTML parser for {os.path.basename(file_path)}")
            sheet_names = [os.path.basename(file_path)]
            read_sheet = lambda _: pd.read_html(file_path)[0]

        for sheet_name in sheet_names:
            try:
                df = read_sheet(sheet_name)
            except Exception as e:
                print(f"⚠️ Skipped sheet {sheet_name} due to error: {e}")
                continue

            if COLS['manifest'] not in df.columns:
                print(f"⚠️ Skipped sheet {sheet_name} — missing Manifest Date")
                continue

            # Output folder by date range
            date_range = pd.to_datetime(df[COLS['manifest']], errors='coerce')
            start_date = date_range.min().strftime('%Y-%m-%d')
            end_date = date_range.max().strftime('%Y-%m-%d')
            folder_name = f"{output_folder}/{start_date}_to_{end_date}"
            os.makedirs(folder_name, exist_ok=True)

            # Remove Duplicates
            before = len(df)
            df = df.drop_duplicates(subset=[COLS['connote']], keep='first')
            after = len(df)
            log.append({'Column': 'Connote #', 'Updated Count': before - after, 'Action': 'Removed Duplicates', 'Sheet': sheet_name})

            # Drop rows with all blank AA–AE
            event_cols = [COLS['AA'], COLS['AB'], COLS['AC'], COLS['AD'], COLS['AE']]
            before = len(df)
            df = df.dropna(subset=event_cols, how='all')
            after = len(df)
            log.append({'Column': 'AA to AE', 'Updated Count': before - after, 'Action': 'Removed blank event rows', 'Sheet': sheet_name})

            # Fill missing tracking data
            count_ab = fill_between(df, COLS['AA'], COLS['AB'], COLS['AC'], tracking_log, sheet_name, file_path)
            log.append({'Column': COLS['AB'], 'Updated Count': count_ab, 'Action': 'Filled missing AB', 'Sheet': sheet_name})

            count_ac = fill_between(df, COLS['AB'], COLS['AC'], COLS['AD'], tracking_log, sheet_name, file_path)
            log.append({'Column': COLS['AC'], 'Updated Count': count_ac, 'Action': 'Filled missing AC', 'Sheet': sheet_name})

            count_ad = fill_between(df, COLS['AC'], COLS['AD'], COLS['AE'], tracking_log, sheet_name, file_path)
            log.append({'Column': COLS['AD'], 'Updated Count': count_ad, 'Action': 'Filled missing AD', 'Sheet': sheet_name})

            # SLA Calculation
            df['SLA Days'] = (pd.to_datetime(df[COLS['AF']], errors='coerce') - pd.to_datetime(df[COLS['AC']], errors='coerce')).dt.days

            def sla_status(days):
                if pd.isna(days):
                    return 'Pending'
                elif days <= 8:
                    return 'Green'
                elif days <= 10:
                    return 'Yellow'
                else:
                    return 'Red'

            df['SLA Status'] = df['SLA Days'].apply(sla_status)

            # Contractor Summary
            summary = df.groupby(COLS['driver']).agg(
                Delivered=('Connote #', lambda x: x[df.loc[x.index, 'SLA Status'].isin(['Green', 'Yellow', 'Red'])].count()),
                Delivered_SLA=('Connote #', lambda x: x[df.loc[x.index, 'SLA Status'] == 'Green'].count()),
                Delivered_Exceeded=('Connote #', lambda x: x[df.loc[x.index, 'SLA Status'] == 'Red'].count()),
                Pending=('Connote #', lambda x: x[df.loc[x.index, 'SLA Status'] == 'Pending'].count()),
                Total=('Connote #', 'count')
            ).reset_index()

            summary['%Delivered SLA'] = (summary['Delivered_SLA'] / summary['Total']).round(2)
            summary['%Pending'] = (summary['Pending'] / summary['Total']).round(2)
            summary['%Delivered'] = (summary['Delivered'] / summary['Total']).round(2)

            # Save output files
            output_file = f"{folder_name}/{sheet_name}_Processed.xlsx"
            df.to_excel(output_file, index=False)
            summary.to_excel(f"{folder_name}/{sheet_name}_Contractor_Summary.xlsx", index=False)

            # Highlight new tracking dates
            wb = load_workbook(output_file)
            ws = wb.active
            for update in tracking_log:
                if update['Sheet'] == sheet_name and os.path.basename(file_path) == update['Source File']:
                    col_letter = chr(ord('A') + list(df.columns).index(update['Updated Column']))
                    row_idx = update['Row Index'] + 2
                    ws[f"{col_letter}{row_idx}"].fill = highlight_fill
            wb.save(output_file)

    # Export logs
    pd.DataFrame(log).to_excel(f"{output_folder}/Tracking_Audit_Log.xlsx", index=False)
    pd.DataFrame(tracking_log).to_excel(f"{output_folder}/Tracking_Update_Log.xlsx", index=False)

    print(f"✅ Processing complete. Reports saved to '{output_folder}'.")

